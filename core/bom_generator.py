# 核心逻辑，处理并生成文件

from typing import Dict, Any, List, Union
import pandas as pd
import json
import openpyxl
import os
import sys
from datetime import datetime
import io


def resource_path(relative_path: str) -> str:
    """
    构建资源文件的完整路径。
    
    Args:
        relative_path (str): 相对于resources目录的路径，如 'category_mapping.json' 或 'templates/上衣模板.xlsx'
        
    Returns:
        str: 完整的资源文件路径
    """
    # 基路径现在是 core 文件夹
    base_path = os.path.dirname(__file__)
    return os.path.join(base_path, 'resources', relative_path)


class BomGenerator:
    """BOM生成器类，用于处理Excel文件并生成BOM表
    
    该类主要用于读取包含产品研发明细信息的Excel文件，
    并提供根据款式编码查找产品信息、生成SKU以及生成完整BOM文件的功能。
    
    支持的操作：
    - 读取Excel文件中的产品明细数据
    - 根据款式编码查找对应的产品信息
    - 返回包含波段、品类、开发颜色等关键信息的结构化数据
    - 根据款式编码、开发颜色和尺码生成对应的SKU列表
    - 基于BOM模板生成完整的Excel BOM文件，支持动态颜色数量
    
    Example:
        >>> generator = BomGenerator('product_details.xlsx')
        >>> info = generator.find_style_info('H5A123416')
        >>> print(info['品类'])  # 输出: 长袖T恤
        >>> skus = generator.generate_skus('H5A413492', '灰色/黑色', ['S', 'M'])
        >>> print(skus[0]['skus']['S'])  # 输出: H5A41349215S
    """
    
    # 类级别常量定义，避免魔术字符串
    SHEET_NAME = '明细表'
    STYLE_CODE_COL = '款式编码'
    WAVE_COL = '波段'
    CATEGORY_COL = '品类'
    DEV_COLOR_COL = '开发颜色'
    
    # BOM模板单元格位置配置
    CELL_CONFIG = {
        # 静态/半静态字段位置
        'timestamp': 'J2',           # 当前时间
        'style_code': 'B3',          # 款式编码
        'order_type': 'F3',          # 订单类型（固定"首单"）
        'product_name_b4': 'B4',     # 品名（B4位置）
        'wave_info': 'F4',           # 波段信息
        'primary_category': 'J4',    # 一级品类信息
        'secondary_category': 'J5',  # 二级品类信息
    }
    
    # 预设颜色块的精确位置映射配置（非均匀布局）
    PRESET_COLOR_BLOCKS = [
        {   # 第1个颜色块的位置映射
            'color_cell': 'A8',     # '下单颜色' 所在的单元格
            'sku_row': 6            # 该颜色对应的 '规格码' 所在的行号
        },
        {   # 第2个颜色块的位置映射
            'color_cell': 'A11',
            'sku_row': 9
        },
        {   # 第3个颜色块的位置映射
            'color_cell': 'A14',
            'sku_row': 12
        }
    ]
    
    def __init__(self, source_path: Union[str, io.BytesIO]) -> None:
        """初始化BomGenerator实例
        
        读取指定的Excel文件或字节流，解析产品明细数据并存储在内存中供后续查询使用。
        Excel文件需要包含"明细表"工作表，且具有特定的表头结构。
        
        Args:
            source_path (Union[str, io.BytesIO]): 源Excel文件的完整路径或字节流对象
            
        Raises:
            FileNotFoundError: 当指定的Excel文件不存在时
            ValueError: 当Excel文件格式不正确或缺少必要工作表时
            
        Note:
            Excel文件应包含复杂的多行表头结构，该方法会自动处理表头解析。
        """
        # 使用resource_path函数获取正确的资源文件路径
        self.template_path = resource_path('bom_template.xlsx')
        self.color_codes_path = resource_path('color_codes.json')
        
        try:
            # 读取指定Excel文件或字节流中的"明细表"Sheet，手动处理复杂表头
            temp_df = pd.read_excel(source_path, sheet_name=self.SHEET_NAME, header=1)
            
            # 使用第0行（现在是DataFrame的第一行）作为列名
            new_columns = temp_df.iloc[0].values
            self.df = temp_df.iloc[1:].copy()
            self.df.columns = new_columns
            
            # 验证必要的列是否存在
            required_columns = [self.STYLE_CODE_COL, self.WAVE_COL, 
                              self.CATEGORY_COL, self.DEV_COLOR_COL]
            missing_columns = [col for col in required_columns if col not in self.df.columns]
            
            if missing_columns:
                raise ValueError(f"Excel文件缺少必要的列: {missing_columns}")
            
            # 加载颜色代码映射表
            try:
                with open(self.color_codes_path, 'r', encoding='utf-8') as f:
                    self.color_codes = json.load(f)
            except FileNotFoundError:
                raise FileNotFoundError(f"错误：颜色代码文件未找到，路径：{self.color_codes_path}")
            except json.JSONDecodeError as e:
                raise ValueError(f"错误：颜色代码文件格式不正确：{str(e)}")
            
            # 加载品类映射
            self.category_mapping_path = resource_path('category_mapping.json')
            try:
                with open(self.category_mapping_path, 'r', encoding='utf-8') as f:
                    self.category_mapping = json.load(f)
            except FileNotFoundError:
                raise FileNotFoundError("错误：品类映射文件 'category_mapping.json' 未找到。")
            except json.JSONDecodeError:
                raise ValueError("错误：品类映射文件 'category_mapping.json' 格式不正确。")
                
        except FileNotFoundError as e:
            if "颜色代码文件" in str(e):
                raise e
            raise FileNotFoundError(f"错误：源文件未找到，路径：{source_path}")
        except Exception as e:
            if "No sheet named" in str(e):
                raise ValueError(f"错误：Excel文件中未找到工作表 '{self.SHEET_NAME}'")
            raise ValueError(f"读取Excel文件时发生错误: {str(e)}")
    
    def find_style_info(self, style_code: str) -> Dict[str, Any]:
        """根据款式编码查找对应的产品样式信息
        
        在已加载的产品数据中搜索指定的款式编码，返回该产品的关键信息。
        返回的信息包括波段、品类和开发颜色等核心产品属性。
        
        Args:
            style_code (str): 要查找的款式编码，如 'H5A123416'
            
        Returns:
            Dict[str, Any]: 包含产品信息的字典，格式如下：
                {
                    '波段': str,        # 产品所属波段，如 '秋四波'
                    '品类': str,        # 产品品类，如 '长袖T恤'
                    '开发颜色': str      # 开发颜色信息，如 '黑色/红色'
                }
                
        Raises:
            ValueError: 当指定的款式编码在数据中不存在时
            
        Example:
            >>> info = generator.find_style_info('H5A123416')
            >>> print(info)
            {'波段': '秋四波', '品类': '长袖T恤', '开发颜色': '黑色/红色'}
        """
        # 查找款式编码匹配的行
        matching_rows = self.df[self.df[self.STYLE_CODE_COL] == style_code]
        
        # 检查是否找到匹配的行
        if matching_rows.empty:
            raise ValueError(f"错误：未在源文件中找到款式编码 '{style_code}'。")
        
        # 获取第一个匹配的行
        row = matching_rows.iloc[0]
        
        # 构建并返回字典
        return {
            self.WAVE_COL: row[self.WAVE_COL],
            self.CATEGORY_COL: row[self.CATEGORY_COL],
            self.DEV_COLOR_COL: row[self.DEV_COLOR_COL]
        }
    
    def get_all_style_codes(self) -> list:
        """
        获取源文件中所有不为空且唯一的款式编码。

        :return: 一个包含所有款式编码字符串的列表。
        """
        if self.STYLE_CODE_COL in self.df.columns:
            return self.df[self.STYLE_CODE_COL].dropna().unique().tolist()
        else:
            raise ValueError(f"错误：源文件中未找到列 '{self.STYLE_CODE_COL}'。")
    
    def generate_skus(self, style_code: str, dev_colors_str: str, sizes: List[str]) -> List[Dict[str, Any]]:
        """根据款式编码、开发颜色和尺码生成SKU列表
        
        根据输入的款式编码、开发颜色字符串和尺码列表，生成对应的SKU信息。
        该方法会解析颜色字符串，查找每个颜色对应的数字代码，然后为每个颜色-尺码
        组合生成唯一的SKU编码。
        
        SKU生成规则：{款式编码}{颜色代码}{尺码}
        例如：H5A41349215S 表示款式H5A413492的灰色(15)S码产品
        
        Args:
            style_code (str): 产品款式编码，如 'H5A413492'
            dev_colors_str (str): 开发颜色字符串，多个颜色用'/'分隔，
                                如 '灰色/黑色/杏色'
            sizes (List[str]): 产品尺码列表，如 ['S', 'M', 'L', 'XL']
            
        Returns:
            List[Dict[str, Any]]: SKU信息列表，每个元素包含以下结构：
                [
                    {
                        'color': str,           # 颜色名称，如 '灰色'
                        'skus': Dict[str, str]  # 尺码到SKU的映射，如 {'S': 'H5A41349215S'}
                    },
                    ...
                ]
                
        Raises:
            ValueError: 当颜色名称在颜色代码字典中找不到时
            
        Example:
            >>> generator = BomGenerator('data.xlsx')
            >>> skus = generator.generate_skus('H5A413492', '灰色/黑色', ['S', 'M'])
            >>> print(skus)
            [
                {
                    'color': '灰色',
                    'skus': {'S': 'H5A41349215S', 'M': 'H5A41349215M'}
                },
                {
                    'color': '黑色', 
                    'skus': {'S': 'H5A41349210S', 'M': 'H5A41349210M'}
                }
            ]
        """
        # 初始化结果列表
        result_list = []
        
        # 根据 '/' 分割颜色字符串
        color_names = [color.strip() for color in dev_colors_str.split('/')]
        
        # 遍历每种颜色
        for color_name in color_names:
            try:
                # 从颜色代码字典中查找对应的代码
                color_code = self.color_codes[color_name]
            except KeyError:
                raise ValueError(f"错误：在颜色代码字典中未找到颜色 '{color_name}'。")
            
            # 创建该颜色的SKU字典
            skus_dict = {}
            
            # 遍历每个尺码
            for size in sizes:
                # 生成SKU字符串
                sku = self._create_sku(style_code, color_code, size)
                # 存入字典
                skus_dict[size] = sku
            
            # 将该颜色的完整信息添加到结果列表
            result_list.append({
                'color': color_name,
                'skus': skus_dict
            })
        
        return result_list
    
    def _create_sku(self, style_code: str, color_code: str, size: str) -> str:
        """创建单个SKU编码
        
        根据款式编码、颜色代码和尺码生成SKU字符串。
        这是一个私有辅助方法，用于保持SKU生成逻辑的一致性。
        
        Args:
            style_code (str): 款式编码
            color_code (str): 颜色数字代码
            size (str): 尺码
            
        Returns:
            str: 生成的SKU编码
            
        Example:
            >>> sku = self._create_sku('H5A413492', '15', 'S')
            >>> print(sku)  # 输出: H5A41349215S
        """
        return f"{style_code}{color_code}{size}"
    
    def generate_bom_file(self, style_code: str, output_dir: str) -> None:
        """生成完整的BOM Excel文件
        
        基于预定义的BOM模板，为指定的款式编码生成包含所有颜色和SKU信息的
        完整Excel BOM文件。该方法支持动态数量的颜色，当颜色超过3种时会
        自动插入新行来容纳额外的颜色信息。
        
        处理流程：
        1. 确保输出目录存在
        2. 加载BOM模板文件
        3. 查找并填充产品基本信息（品名等）
        4. 生成所有颜色的SKU信息
        5. 如果颜色数量超过3种，动态插入新行
        6. 填充所有颜色和SKU信息到对应位置
        7. 保存为新的Excel文件
        
        Args:
            style_code (str): 产品款式编码，如 'H5A123416'
            output_dir (str): 输出目录的完整路径。如果目录不存在会自动创建
            
        Raises:
            ValueError: 当款式编码在源数据中不存在时
            FileNotFoundError: 当BOM模板文件不存在时
            PermissionError: 当无法创建输出目录或写入文件时
            
        Example:
            >>> generator = BomGenerator('source.xlsx')
            >>> generator.generate_bom_file('H5A123416', './output')
            # 将在 ./output/ 目录下生成 H5A123416.xlsx 文件
            
        Note:
            - 生成的文件名格式为: {款式编码}.xlsx
            - 支持任意数量的颜色，超过3种会自动扩展表格行数
            - 品名格式为: HECO{波段}{品类}{款式编码}
            - SKU生成规则: {款式编码}{颜色代码}{尺码}
        """
        # 1. 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)
        
        # 2. 加载模板文件
        try:
            workbook = openpyxl.load_workbook(self.template_path)
            sheet = workbook.active
        except FileNotFoundError:
            raise FileNotFoundError(f"错误：BOM模板文件未找到，路径：{self.template_path}")
        
        # 3. 获取产品基本信息
        style_info = self.find_style_info(style_code)
        
        # 4. 生成品名（HECO + 波段 + 品类 + 款式编码）
        product_name = f"HECO{style_info[self.WAVE_COL]}{style_info[self.CATEGORY_COL]}{style_code}"
        
        # 5. 填充静态/半静态字段内容
        config = self.CELL_CONFIG
        
        # 当前时间格式化为 YYYY/MM/DD HH:MM
        current_time = datetime.now().strftime("%Y/%m/%d %H:%M")
        self._write_to_cell(sheet, config['timestamp'], current_time)
        
        # 款式编码
        self._write_to_cell(sheet, config['style_code'], style_code)
        
        # 固定写入 "首单"
        self._write_to_cell(sheet, config['order_type'], "首单")
        
        # 品名（B4位置）
        self._write_to_cell(sheet, config['product_name_b4'], product_name)
        
        # 波段信息
        self._write_to_cell(sheet, config['wave_info'], style_info[self.WAVE_COL])
        
        # 6. 生成SKU列表
        dev_colors = style_info[self.DEV_COLOR_COL]
        sizes = ['S', 'M', 'L', 'XL']
        sku_list = self.generate_skus(style_code, dev_colors, sizes)
        
        # 7. 使用精确位置映射填充颜色和SKU信息（最多处理前3个颜色）
        for i, color_info in enumerate(sku_list):
            if i >= len(self.PRESET_COLOR_BLOCKS):
                break  # 暂时不处理超过3个的颜色
            
            try:
                block_config = self.PRESET_COLOR_BLOCKS[i]
                color_cell_addr = block_config['color_cell']
                sku_target_row = block_config['sku_row']
                
                # 1. 写入颜色名称
                self._write_to_cell(sheet, color_cell_addr, color_info['color'])
                
                # 2. 写入规格码 (SKU) - 使用精确的行号和列号
                # SKU从B列开始，行号由 sku_target_row 决定
                sheet.cell(row=sku_target_row, column=2).value = color_info['skus']['S']  # B列
                sheet.cell(row=sku_target_row, column=3).value = color_info['skus']['M']  # C列
                sheet.cell(row=sku_target_row, column=4).value = color_info['skus']['L']  # D列
                sheet.cell(row=sku_target_row, column=5).value = color_info['skus']['XL'] # E列
                
            except Exception as e:
                # 提供详细的错误信息
                raise ValueError(f"填充第{i+1}个颜色块时出错 (颜色: {color_info['color']}): {str(e)}")
        
        # 9. 保存文件
        output_file_path = os.path.join(output_dir, f"{style_code}.xlsx")
        try:
            workbook.save(output_file_path)
        except PermissionError:
            raise PermissionError(f"错误：无法保存文件到 {output_file_path}，请检查目录权限。")
    
    def generate_bom_file_to_buffer(self, style_code: str) -> bytes:
        """
        生成单个BOM Excel文件，并将其作为字节流返回。
        
        基于动态选择的BOM模板，为指定的款式编码生成包含所有颜色和SKU信息的
        完整Excel BOM文件，并返回文件的字节内容而不是保存到磁盘。
        
        Args:
            style_code (str): 产品款式编码，如 'H5A123416'
            
        Returns:
            bytes: Excel文件的字节内容
            
        Raises:
            ValueError: 当款式编码在源数据中不存在时或未定义的品类时
            FileNotFoundError: 当BOM模板文件不存在时
            
        Example:
            >>> generator = BomGenerator('source.xlsx')
            >>> excel_bytes = generator.generate_bom_file_to_buffer('H5A123416')
            >>> len(excel_bytes)  # 返回文件字节长度
            
        Note:
            - 支持任意数量的颜色，超过3种会自动扩展表格行数
            - 品名格式为: HECO{波段}{品类}{款式编码}
            - SKU生成规则: {款式编码}{颜色代码}{尺码}
            - 动态选择模板：根据二级品类映射到一级品类，选择对应模板
        """
        # 1. 获取产品基本信息
        style_info = self.find_style_info(style_code)
        
        # 2. 动态模板选择逻辑
        # a. 获取当前处理款式的二级品类（secondary_category）
        secondary_category = style_info[self.CATEGORY_COL]
        
        # b. 使用self.category_mapping字典，根据secondary_category查找对应的一级品类（primary_category）
        if secondary_category not in self.category_mapping:
            raise ValueError(f"错误：未定义的品类 '{secondary_category}'，请在category_mapping.json中配置。")
        
        primary_category = self.category_mapping[secondary_category]
        
        # c. 根据primary_category构建模板文件的路径
        template_path = resource_path(f'templates/{primary_category}模板.xlsx')
        
        # d. 加载这个动态确定的模板文件
        try:
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook.active
        except FileNotFoundError:
            raise FileNotFoundError(f"错误：BOM模板文件未找到，路径：{template_path}")
        
        # 3. 填充一级品类和二级品类
        sheet['J4'] = primary_category
        sheet['J5'] = secondary_category
        
        # 4. 生成品名（HECO + 波段 + 品类 + 款式编码）
        product_name = f"HECO{style_info[self.WAVE_COL]}{style_info[self.CATEGORY_COL]}{style_code}"
        
        # 4. 填充静态/半静态字段内容
        config = self.CELL_CONFIG
        
        # 当前时间格式化为 YYYY/MM/DD HH:MM
        current_time = datetime.now().strftime("%Y/%m/%d %H:%M")
        self._write_to_cell(sheet, config['timestamp'], current_time)
        
        # 款式编码
        self._write_to_cell(sheet, config['style_code'], style_code)
        
        # 固定写入 "首单"
        self._write_to_cell(sheet, config['order_type'], "首单")
        
        # 品名（B4位置）
        self._write_to_cell(sheet, config['product_name_b4'], product_name)
        
        # 波段信息
        self._write_to_cell(sheet, config['wave_info'], style_info[self.WAVE_COL])
        
        # 5. 生成SKU列表
        dev_colors = style_info[self.DEV_COLOR_COL]
        sizes = ['S', 'M', 'L', 'XL']
        sku_list = self.generate_skus(style_code, dev_colors, sizes)
        
        # 6. 使用精确位置映射填充颜色和SKU信息（最多处理前3个颜色）
        for i, color_info in enumerate(sku_list):
            if i >= len(self.PRESET_COLOR_BLOCKS):
                break  # 暂时不处理超过3个的颜色
            
            try:
                block_config = self.PRESET_COLOR_BLOCKS[i]
                color_cell_addr = block_config['color_cell']
                sku_target_row = block_config['sku_row']
                
                # 1. 写入颜色名称
                self._write_to_cell(sheet, color_cell_addr, color_info['color'])
                
                # 2. 写入规格码 (SKU) - 使用精确的行号和列号
                # SKU从B列开始，行号由 sku_target_row 决定
                sheet.cell(row=sku_target_row, column=2).value = color_info['skus']['S']  # B列
                sheet.cell(row=sku_target_row, column=3).value = color_info['skus']['M']  # C列
                sheet.cell(row=sku_target_row, column=4).value = color_info['skus']['L']  # D列
                sheet.cell(row=sku_target_row, column=5).value = color_info['skus']['XL'] # E列
                
            except Exception as e:
                # 提供详细的错误信息
                raise ValueError(f"填充第{i+1}个颜色块时出错 (颜色: {color_info['color']}): {str(e)}")
        
        # 7. 将工作簿保存在内存中的字节流中
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)  # 将指针移回开头
        
        return buffer.getvalue()
    
    def _write_to_cell(self, sheet, cell_address: str, value: str) -> None:
        """向Excel单元格写入值，处理合并单元格情况
        
        Args:
            sheet: openpyxl工作表对象
            cell_address (str): 单元格地址，如 'C4'
            value (str): 要写入的值
        """
        target_cell = sheet[cell_address]
        if target_cell.__class__.__name__ == 'MergedCell':
            # 找到包含目标单元格的合并区域的主单元格
            for range_ in sheet.merged_cells.ranges:
                if cell_address in range_:
                    # 写入到合并区域的左上角单元格
                    main_cell = sheet.cell(row=range_.min_row, column=range_.min_col)
                    main_cell.value = value
                    break
            else:
                # 如果找不到合并区域，直接写入目标单元格
                target_cell.value = value
        else:
            target_cell.value = value
    
    def _insert_additional_rows(self, sheet, color_count: int) -> None:
        """当颜色数量超过3种时，动态插入新行
        
        Args:
            sheet: openpyxl工作表对象
            color_count (int): 颜色总数
        """
        if color_count <= 3:
            return
        
        # 计算需要插入的行数：每超过一种颜色需要插入3行（颜色行+SKU行+空行）
        additional_colors = color_count - 3
        rows_to_insert = additional_colors * 3
        
        # 在指定位置插入新行
        insert_row = self.CELL_CONFIG['insert_before_row']
        sheet.insert_rows(insert_row, rows_to_insert)
    
    def _fill_color_and_sku_data(self, sheet, sku_list: List[Dict[str, Any]]) -> None:
        """填充所有颜色和SKU信息到工作表
        
        Args:
            sheet: openpyxl工作表对象
            sku_list (List[Dict[str, Any]]): SKU信息列表
        """
        config = self.CELL_CONFIG
        color_column = config['color_column']
        sku_columns = config['sku_columns']
        
        for i, color_data in enumerate(sku_list):
            # 计算当前颜色块的行位置
            color_row = config['color_start_row'] + (i * config['rows_per_color'])
            sku_row = config['sku_start_row'] + (i * config['rows_per_color'])
            
            # 填充颜色名称 - 使用_write_to_cell处理合并单元格
            color_cell_address = f"{color_column}{color_row}"
            self._write_to_cell(sheet, color_cell_address, color_data['color'])
            
            # 填充SKU信息 - 使用_write_to_cell处理合并单元格
            for size, sku in color_data['skus'].items():
                if size in sku_columns:
                    column = sku_columns[size]
                    sku_cell_address = f"{column}{sku_row}"
                    self._write_to_cell(sheet, sku_cell_address, sku)
    
    def _fill_color_and_sku_data_precise(self, sheet, sku_list: List[Dict[str, Any]]) -> None:
        """精确填充所有颜色和SKU信息到工作表，完全匹配标准模板布局
        
        Args:
            sheet: openpyxl工作表对象
            sku_list (List[Dict[str, Any]]): SKU信息列表
        """
        config = self.CELL_CONFIG
        color_column = config['color_column']
        color_merge_end = config['color_merge_end']
        sku_columns = config['sku_columns']
        color_rows = config['color_rows']  # [8, 11, 13] - 预定义的颜色行号
        
        for i, color_data in enumerate(sku_list):
            # 确定当前颜色的行号
            if i < len(color_rows):
                # 使用预定义的行号
                color_row = color_rows[i]
            else:
                # 超过预定义数量，需要计算额外插入的行号
                # 基于最后一个预定义行号，每个额外颜色增加3行间距
                extra_colors = i - len(color_rows) + 1
                color_row = color_rows[-1] + (extra_colors * 3)
            
            # 规格码行号是颜色行号的下一行
            sku_row = color_row + 1
            
            # 填充颜色名称并合并单元格
            color_cell_address = f"{color_column}{color_row}"
            self._write_to_cell(sheet, color_cell_address, color_data['color'])
            
            # 合并单元格：从B列到G列
            merge_range = f"{color_column}{color_row}:{color_merge_end}{color_row}"
            try:
                sheet.merge_cells(merge_range)
            except ValueError:
                # 如果单元格已经合并，忽略错误
                pass
            
            # 填充SKU信息到规格码行
            for size, sku in color_data['skus'].items():
                if size in sku_columns:
                    column = sku_columns[size]
                    sku_cell_address = f"{column}{sku_row}"
                    self._write_to_cell(sheet, sku_cell_address, sku)
